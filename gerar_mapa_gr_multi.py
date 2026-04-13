#!/usr/bin/env python3
"""
MAPA-GR v2.0 — Gerador de dataset consolidado por grande área.

Lê todos os arquivos CPC/ENADE/IDD da pasta dados_inep/ e produz 10 arquivos JSON
em docs/dados/ e dados/ (raiz), um por grande área CAPES/CNPq + metadata.json.

Fontes:
  - CPC_2017.xlsx (Licenciaturas pré-COVID + Eng)
  - CPC_2021.xlsx (Licenciaturas pós-COVID)
  - CPC_2022.xlsx (Sociais Aplicadas)
  - CPC_2023.xlsx (Saúde/Eng/Agro)
  - conceito_enade_2018.xlsx + IDD_2018.xlsx (Sociais Aplicadas 2018, parcial)

Saída: docs/dados/{grande_area}.json (minificado)
"""

import openpyxl
import json
import os
from collections import defaultdict
from datetime import datetime

BASE = os.path.dirname(os.path.abspath(__file__))
IN_DIR = os.path.join(BASE, 'dados_inep')

# ===== Mapeamento Área INEP → Grande Área =====
AREA_TO_GROUP = {
    # Exatas e da Terra
    'FÍSICA (LICENCIATURA)': 'exatas',
    'FÍSICA (BACHARELADO)': 'exatas',
    'MATEMÁTICA (LICENCIATURA)': 'exatas',
    'MATEMÁTICA (BACHARELADO)': 'exatas',
    'QUÍMICA (LICENCIATURA)': 'exatas',
    'QUÍMICA (BACHARELADO)': 'exatas',
    'CIÊNCIA DA COMPUTAÇÃO (BACHARELADO)': 'exatas',
    'CIÊNCIA DA COMPUTAÇÃO (LICENCIATURA)': 'exatas',
    'SISTEMAS DE INFORMAÇÃO': 'exatas',
    # Biológicas
    'CIÊNCIAS BIOLÓGICAS (LICENCIATURA)': 'biologicas',
    'CIÊNCIAS BIOLÓGICAS (BACHARELADO)': 'biologicas',
    'BIOMEDICINA': 'biologicas',
    # Engenharias
    'ENGENHARIA': 'engenharias',
    'ENGENHARIA AMBIENTAL': 'engenharias',
    'ENGENHARIA CIVIL': 'engenharias',
    'ENGENHARIA DA COMPUTAÇÃO': 'engenharias',
    'ENGENHARIA DE ALIMENTOS': 'engenharias',
    'ENGENHARIA DE COMPUTAÇÃO I': 'engenharias',
    'ENGENHARIA DE CONTROLE E AUTOMAÇÃO': 'engenharias',
    'ENGENHARIA DE PRODUÇÃO': 'engenharias',
    'ENGENHARIA ELÉTRICA': 'engenharias',
    'ENGENHARIA FLORESTAL': 'engenharias',
    'ENGENHARIA MECÂNICA': 'engenharias',
    'ENGENHARIA QUÍMICA': 'engenharias',
    'ARQUITETURA E URBANISMO': 'engenharias',
    # Saúde
    'ENFERMAGEM': 'saude',
    'FARMÁCIA': 'saude',
    'FISIOTERAPIA': 'saude',
    'FONOAUDIOLOGIA': 'saude',
    'MEDICINA': 'saude',
    'NUTRIÇÃO': 'saude',
    'ODONTOLOGIA': 'saude',
    'PSICOLOGIA': 'saude',
    'EDUCAÇÃO FÍSICA (LICENCIATURA)': 'saude',
    'EDUCAÇÃO FÍSICA (BACHARELADO)': 'saude',
    # Agrárias
    'AGRONOMIA': 'agrarias',
    'ZOOTECNIA': 'agrarias',
    'MEDICINA VETERINÁRIA': 'agrarias',
    # Sociais Aplicadas
    'ADMINISTRAÇÃO': 'sociais',
    'ADMINISTRAÇÃO PÚBLICA': 'sociais',
    'DIREITO': 'sociais',
    'CIÊNCIAS CONTÁBEIS': 'sociais',
    'CIÊNCIAS ECONÔMICAS': 'sociais',
    'COMUNICAÇÃO SOCIAL - JORNALISMO': 'sociais',
    'JORNALISMO': 'sociais',
    'COMUNICAÇÃO SOCIAL - PUBLICIDADE E PROPAGANDA': 'sociais',
    'PUBLICIDADE E PROPAGANDA': 'sociais',
    'SERVIÇO SOCIAL': 'sociais',
    'RELAÇÕES INTERNACIONAIS': 'sociais',
    'TURISMO': 'sociais',
    'DESIGN': 'sociais',
    'SECRETARIADO EXECUTIVO': 'sociais',
    # Humanas
    'PEDAGOGIA (LICENCIATURA)': 'humanas',
    'HISTÓRIA (LICENCIATURA)': 'humanas',
    'HISTÓRIA (BACHARELADO)': 'humanas',
    'GEOGRAFIA (LICENCIATURA)': 'humanas',
    'GEOGRAFIA (BACHARELADO)': 'humanas',
    'FILOSOFIA (LICENCIATURA)': 'humanas',
    'FILOSOFIA (BACHARELADO)': 'humanas',
    'CIÊNCIAS SOCIAIS (LICENCIATURA)': 'humanas',
    'CIÊNCIAS SOCIAIS (BACHARELADO)': 'humanas',
    'TEOLOGIA': 'humanas',
    # Letras e Artes
    'LETRAS-PORTUGUÊS (LICENCIATURA)': 'letras',
    'LETRAS-PORTUGUÊS (BACHARELADO)': 'letras',
    'LETRAS-PORTUGUÊS E ESPANHOL (LICENCIATURA)': 'letras',
    'LETRAS-PORTUGUÊS E INGLÊS (LICENCIATURA)': 'letras',
    'Letras - Inglês': 'letras',
    'MÚSICA (LICENCIATURA)': 'letras',
    'ARTES VISUAIS (LICENCIATURA)': 'letras',
}

GROUP_LABELS = {
    'exatas': 'Ciências Exatas e da Terra',
    'biologicas': 'Ciências Biológicas',
    'engenharias': 'Engenharias',
    'saude': 'Ciências da Saúde',
    'agrarias': 'Ciências Agrárias',
    'sociais': 'Ciências Sociais Aplicadas',
    'humanas': 'Ciências Humanas',
    'letras': 'Linguística, Letras e Artes',
    'tecnologos': 'Cursos Tecnólogos',
}

GROUP_ICONS = {
    'exatas': '🔬', 'biologicas': '🌿', 'engenharias': '⚙️',
    'saude': '⚕️', 'agrarias': '🌾', 'sociais': '💼',
    'humanas': '📚', 'letras': '🎨', 'tecnologos': '🛠️',
}

GROUP_ORDER = ['exatas', 'biologicas', 'engenharias', 'saude', 'agrarias',
               'sociais', 'humanas', 'letras', 'tecnologos']


def group_of(area):
    if area in AREA_TO_GROUP:
        return AREA_TO_GROUP[area]
    if 'TECNOLOGIA EM' in area.upper() or 'TECNOLÓGICO' in area.upper():
        return 'tecnologos'
    return None


def fnum(v):
    if v is None or v in ('', '-', 'SC', 'NS'):
        return None
    try:
        return round(float(str(v).replace(',', '.')), 3)
    except Exception:
        return None


def iint(v):
    if v is None or v == '':
        return None
    try:
        return int(float(str(v).replace(',', '.')))
    except Exception:
        return None


def s(v):
    return '' if v is None else str(v).strip()


def clean(d):
    return {k: v for k, v in d.items() if v not in (None, '')}


# ===== Extraction functions =====

def extract_cpc_2017(path):
    """CPC 2017: Licenciaturas + alguns Bacharelados + Engenharias pré-COVID."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['PORTAL_CPC_2017']
    out = []
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or not r or not r[0]:
            continue
        area = s(r[2])
        grau = ('Bacharelado' if '(BACHARELADO)' in area.upper()
                else 'Licenciatura' if '(LICENCIATURA)' in area.upper() else 'Outro')
        d = {
            'y': 2017, 'ar': area, 'g': grau, 'ci': iint(r[3]), 'ie': s(r[4]), 'sg': s(r[5]),
            'o': s(r[6]), 'ct': s(r[7]), 'cc': iint(r[8]), 'm': s(r[9]), 'mu': s(r[11]), 'u': s(r[12]),
            'ni': iint(r[13]), 'np': iint(r[14]), 'fg': fnum(r[15]), 'ce': fnum(r[16]),
            'e': fnum(r[17]), 'ib': fnum(r[20]), 'ip': fnum(r[21]), 'op': fnum(r[23]),
            'nf': fnum(r[25]), 'of': fnum(r[27]), 'nd': iint(r[28]),
            'ms': fnum(r[30]), 'dr': fnum(r[32]), 'rg': fnum(r[34]),
            'pc': fnum(r[35]), 'pf': iint(r[36]),
        }
        out.append(clean(d))
    return out


def extract_cpc_2021(path):
    """CPC 2021: Licenciaturas pós-COVID."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['CPC2021']
    out = []
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or not r or not r[0]:
            continue
        d = {
            'y': 2021, 'ar': s(r[2]), 'g': s(r[3]), 'ci': iint(r[4]), 'ie': s(r[5]), 'sg': s(r[6]),
            'o': s(r[7]), 'ct': s(r[8]), 'cc': iint(r[9]), 'm': s(r[10]), 'mu': s(r[12]), 'u': s(r[13]),
            'ni': iint(r[14]), 'np': iint(r[15]), 'fg': fnum(r[16]), 'ce': fnum(r[18]),
            'e': fnum(r[20]), 'ib': fnum(r[23]), 'ip': fnum(r[24]), 'op': fnum(r[26]),
            'nf': fnum(r[28]), 'of': fnum(r[30]),
            'ms': fnum(r[32]), 'dr': fnum(r[34]), 'rg': fnum(r[36]),
            'pc': fnum(r[37]), 'pf': iint(r[38]),
        }
        out.append(clean(d))
    return out


def extract_cpc_2022(path):
    """CPC 2022: Sociais Aplicadas pós-COVID."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['CPC 2022']
    out = []
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or not r or not r[0]:
            continue
        d = {
            'y': 2022, 'ar': s(r[8]), 'ci': iint(r[1]), 'ie': s(r[2]), 'sg': s(r[3]),
            'o': s(r[4]), 'ct': s(r[5]), 'cc': iint(r[6]), 'm': s(r[9]), 'mu': s(r[11]), 'u': s(r[12]),
            'ni': iint(r[13]), 'np': iint(r[14]), 'fg': fnum(r[15]), 'ce': fnum(r[17]),
            'e': fnum(r[19]), 'ib': fnum(r[22]), 'ip': fnum(r[23]), 'op': fnum(r[25]),
            'nf': fnum(r[27]), 'of': fnum(r[29]),
            'ms': fnum(r[31]), 'dr': fnum(r[33]), 'rg': fnum(r[35]),
            'pc': fnum(r[36]), 'pf': iint(r[37]),
        }
        out.append(clean(d))
    return out


def extract_cpc_2023(path):
    """CPC 2023: Saúde + Engenharias pós-COVID."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['CPC_2023']
    out = []
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or not r or not r[0]:
            continue
        d = {
            'y': 2023, 'ar': s(r[8]), 'ci': iint(r[1]), 'ie': s(r[2]), 'sg': s(r[3]),
            'o': s(r[4]), 'ct': s(r[5]), 'cc': iint(r[6]), 'm': s(r[9]), 'mu': s(r[11]), 'u': s(r[12]),
            'ni': iint(r[13]), 'np': iint(r[14]), 'fg': fnum(r[15]), 'ce': fnum(r[17]),
            'e': fnum(r[19]), 'ib': fnum(r[22]), 'ip': fnum(r[23]), 'op': fnum(r[25]),
            'nf': fnum(r[27]), 'of': fnum(r[29]),
            'ms': fnum(r[31]), 'dr': fnum(r[33]), 'rg': fnum(r[35]),
            'pc': fnum(r[36]), 'pf': iint(r[37]),
        }
        out.append(clean(d))
    return out


def extract_enade_idd_2018(enade_path, idd_path):
    """2018: só Conceito ENADE + IDD (CPC completo não publicado)."""
    wb_e = openpyxl.load_workbook(enade_path, read_only=True, data_only=True)
    ws_e = wb_e['Conceito_Enade_2018']
    cursos = {}
    for i, r in enumerate(ws_e.iter_rows(values_only=True)):
        if i == 0 or not r:
            continue
        cc = iint(r[8])
        if cc is None:
            continue
        cursos[cc] = {
            'y': 2018, 'ar': s(r[2]), 'ci': iint(r[3]), 'ie': s(r[4]), 'sg': s(r[5]),
            'o': s(r[6]), 'ct': s(r[7]), 'cc': cc, 'm': s(r[9]), 'mu': s(r[11]), 'u': s(r[12]),
            'ni': iint(r[13]), 'np': iint(r[14]), 'fg': fnum(r[15]), 'ce': fnum(r[17]),
            'e': fnum(r[19]), 'pf': iint(r[20]),  # Faixa ENADE (1-5) como proxy do CPC
            '_partial': 1,  # flag: indicadores parciais (sem CPC completo)
        }
    wb_i = openpyxl.load_workbook(idd_path, read_only=True, data_only=True)
    ws_i = wb_i['IDD_2018']
    for i, r in enumerate(ws_i.iter_rows(values_only=True)):
        if i == 0 or not r:
            continue
        cc = iint(r[8])
        if cc is None or cc not in cursos:
            continue
        cursos[cc]['ib'] = fnum(r[17])
        cursos[cc]['ip'] = fnum(r[18])
    return [clean(d) for d in cursos.values()]


# ===== Main =====

def main():
    print(f'Lendo arquivos de {IN_DIR}')
    d17 = extract_cpc_2017(os.path.join(IN_DIR, 'CPC_2017.xlsx'))
    print(f'  CPC 2017: {len(d17)} cursos')
    d18 = extract_enade_idd_2018(
        os.path.join(IN_DIR, 'conceito_enade_2018.xlsx'),
        os.path.join(IN_DIR, 'IDD_2018.xlsx'))
    print(f'  ENADE+IDD 2018: {len(d18)} cursos (parcial, sem CPC composto)')
    d21 = extract_cpc_2021(os.path.join(IN_DIR, 'CPC_2021.xlsx'))
    print(f'  CPC 2021: {len(d21)} cursos')
    d22 = extract_cpc_2022(os.path.join(IN_DIR, 'CPC_2022.xlsx'))
    print(f'  CPC 2022: {len(d22)} cursos')
    d23 = extract_cpc_2023(os.path.join(IN_DIR, 'CPC_2023.xlsx'))
    print(f'  CPC 2023: {len(d23)} cursos')

    all_courses = d17 + d18 + d21 + d22 + d23
    print(f'\nTotal cursos consolidados: {len(all_courses)}')

    # Agrupar por grande área
    groups = defaultdict(list)
    areas_per_group = defaultdict(set)
    cycles_per_group = defaultdict(set)
    ies_per_group = defaultdict(set)

    for c in all_courses:
        g = group_of(c.get('ar', ''))
        if g is None:
            continue
        groups[g].append(c)
        areas_per_group[g].add(c.get('ar', ''))
        cycles_per_group[g].add(c.get('y'))
        if c.get('sg'):
            ies_per_group[g].add(c.get('sg'))

    # Escrever cada grande área
    out_dir_docs = os.path.join(BASE, 'docs', 'dados')
    os.makedirs(out_dir_docs, exist_ok=True)

    metadata = {
        'versao': '2.0',
        'gerado_em': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'fonte': 'INEP — Dados Abertos (CPC/ENADE/IDD)',
        'portal': 'https://www.gov.br/inep/pt-br/acesso-a-informacao/dados-abertos',
        'total_cursos': sum(len(lst) for lst in groups.values()),
        'grandes_areas': {},
    }

    print(f'\n== Escrevendo JSONs ==')
    for g in GROUP_ORDER:
        lst = groups.get(g, [])
        # Ordenar por CPC contínuo desc para facilitar inspeção
        lst.sort(key=lambda c: -(c.get('pc') or 0))
        data = {
            'grande_area': g,
            'label': GROUP_LABELS[g],
            'icon': GROUP_ICONS[g],
            'n_cursos': len(lst),
            'n_ies': len(ies_per_group[g]),
            'ciclos': sorted(cycles_per_group[g]),
            'areas': sorted(areas_per_group[g]),
            'cursos': lst,
        }
        out_path = os.path.join(out_dir_docs, f'{g}.json')
        with open(out_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, separators=(',', ':'))
        size = os.path.getsize(out_path)
        metadata['grandes_areas'][g] = {
            'label': GROUP_LABELS[g],
            'icon': GROUP_ICONS[g],
            'n_cursos': len(lst),
            'n_ies': len(ies_per_group[g]),
            'n_areas': len(areas_per_group[g]),
            'ciclos': sorted(cycles_per_group[g]),
            'arquivo': f'dados/{g}.json',
            'tamanho_bytes': size,
        }
        print(f'  {GROUP_ICONS[g]} {g:14}: {len(lst):6} cursos | {size/1024:6.0f} KB | ciclos {sorted(cycles_per_group[g])}')

    # metadata.json
    meta_path = os.path.join(out_dir_docs, 'metadata.json')
    with open(meta_path, 'w', encoding='utf-8') as f:
        json.dump(metadata, f, ensure_ascii=False, separators=(',', ':'))
    print(f'\n  metadata.json: {os.path.getsize(meta_path)} bytes')

    total_size = sum(m['tamanho_bytes'] for m in metadata['grandes_areas'].values())
    print(f'\nTotal docs/dados/: {total_size/1024/1024:.2f} MB')
    print(f'Total cursos: {metadata["total_cursos"]}')


if __name__ == '__main__':
    main()
