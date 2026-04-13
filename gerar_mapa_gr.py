#!/usr/bin/env python3
"""
MAPA-GR — Gerador de dataset consolidado de cursos de graduação em Física
a partir dos arquivos XLSX dos Indicadores de Qualidade da Educação Superior (INEP).

Fontes lidas da pasta dados_inep/ (não commitada):
  - CPC_2017.xlsx (ciclo Licenciaturas + Bacharelado)
  - CPC_2021.xlsx (ciclo Licenciaturas)
  - IGC_2017.xlsx e IGC_2021.xlsx (conceito institucional da IES)

Escopo v1.0: apenas área FÍSICA (Licenciatura + Bacharelado).
O script gera docs/dados_inep.json (para a PWA) e dados_inep.json (raiz, standalone).
"""

import openpyxl
import json
import os
import gzip
from collections import defaultdict

BASE = os.path.dirname(os.path.abspath(__file__))
IN_DIR = os.path.join(BASE, 'dados_inep')

# ---------- utilitários ----------

def fnum(v):
    """Converte para float ou None."""
    if v is None or v == '' or v == '-':
        return None
    s = str(v).strip()
    if s in ('SC', 'NS', '—', 'N/A'):
        return None
    try:
        return round(float(s.replace(',', '.')), 3)
    except Exception:
        return None

def iint(v):
    if v is None or v == '':
        return None
    try:
        return int(float(str(v).replace(',', '.')))
    except Exception:
        return None

def s(v):
    if v is None:
        return ''
    return str(v).strip()

# ---------- extração CPC 2017 ----------

def extract_cpc_2017(path):
    """CPC 2017 — layout com 37 colunas, sem coluna Grau (inferido da área)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['PORTAL_CPC_2017']
    out = []
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or not r or not r[0]:
            continue
        area = s(r[2])
        if 'FÍSICA' not in area.upper() or 'EDUCAÇÃO' in area.upper():
            continue
        if '(BACHARELADO)' in area.upper():
            grau = 'Bacharelado'
        elif '(LICENCIATURA)' in area.upper():
            grau = 'Licenciatura'
        else:
            continue
        out.append({
            'y': 2017,
            'ar': area,
            'g': grau,
            'ci': iint(r[3]),
            'ie': s(r[4]),
            'sg': s(r[5]),
            'o': s(r[6]),
            'ct': s(r[7]),
            'cc': iint(r[8]),
            'm': s(r[9]),
            'mu': s(r[11]),
            'u': s(r[12]),
            'ni': iint(r[13]),
            'np': iint(r[14]),
            'fg': fnum(r[15]),
            'ce': fnum(r[16]),
            'e':  fnum(r[17]),
            'ib': fnum(r[20]),
            'ip': fnum(r[21]),
            'op': fnum(r[23]),
            'nf': fnum(r[25]),
            'of': fnum(r[27]),
            'nd': iint(r[28]),
            'ms': fnum(r[30]),
            'dr': fnum(r[32]),
            'rg': fnum(r[34]),
            'pc': fnum(r[35]),
            'pf': iint(r[36]),
        })
    return out

# ---------- extração CPC 2021 ----------

def extract_cpc_2021(path):
    """CPC 2021 — layout com 40 colunas, coluna explícita Grau acadêmico (idx 3)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb['CPC2021']
    out = []
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or not r or not r[0]:
            continue
        area = s(r[2])
        if 'FÍSICA' not in area.upper() or 'EDUCAÇÃO' in area.upper():
            continue
        out.append({
            'y': 2021,
            'ar': area,
            'g': s(r[3]),
            'ci': iint(r[4]),
            'ie': s(r[5]),
            'sg': s(r[6]),
            'o': s(r[7]),
            'ct': s(r[8]),
            'cc': iint(r[9]),
            'm': s(r[10]),
            'mu': s(r[12]),
            'u': s(r[13]),
            'ni': iint(r[14]),
            'np': iint(r[15]),
            'fg': fnum(r[16]),
            'ce': fnum(r[18]),
            'e':  fnum(r[20]),
            'ib': fnum(r[23]),
            'ip': fnum(r[24]),
            'op': fnum(r[26]),
            'nf': fnum(r[28]),
            'of': fnum(r[30]),
            'ms': fnum(r[32]),
            'dr': fnum(r[34]),
            'rg': fnum(r[36]),
            'pc': fnum(r[37]),
            'pf': iint(r[38]),
        })
    return out

# ---------- extração IGC (IES-level) ----------

def extract_igc(path, year, sheet_hint):
    """IGC é da IES, não do curso. Retorna dict cd_ies -> igc_faixa."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = None
    for sn in wb.sheetnames:
        if sheet_hint.lower() in sn.lower():
            ws = wb[sn]
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]
    # try to find column "Código da IES" and "IGC Faixa"
    headers = None
    for row in ws.iter_rows(max_row=1, values_only=True):
        headers = row
        break
    if not headers:
        return {}
    col_cd = None
    col_igc_faixa = None
    col_igc_cont = None
    for j, h in enumerate(headers):
        if not h:
            continue
        hl = str(h).lower().replace('á', 'a').replace('ó', 'o').replace('ç', 'c')
        if 'codigo' in hl and 'ies' in hl and col_cd is None:
            col_cd = j
        if 'igc' in hl and ('faixa' in hl or '(faixa)' in hl) and col_igc_faixa is None:
            col_igc_faixa = j
        if 'igc' in hl and ('contin' in hl or 'continuo' in hl or '(continuo)' in hl) and col_igc_cont is None:
            col_igc_cont = j
    if col_cd is None:
        return {}
    res = {}
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if not r or len(r) <= col_cd:
            continue
        cd = iint(r[col_cd])
        if cd is None:
            continue
        f = iint(r[col_igc_faixa]) if col_igc_faixa is not None and col_igc_faixa < len(r) else None
        c = fnum(r[col_igc_cont]) if col_igc_cont is not None and col_igc_cont < len(r) else None
        if cd not in res or (res[cd].get('igc_f') is None and f is not None):
            res[cd] = {'igc_f': f, 'igc_c': c}
    return res

# ---------- agregações ----------

def aggregate_by_ies(cursos):
    """Para cada IES, calcular CPC médio, nº de cursos, melhor faixa."""
    by = defaultdict(list)
    for c in cursos:
        if c.get('pc') is not None:
            by[c['sg']].append(c)
    res = {}
    for sig, lst in by.items():
        cpcs = [c['pc'] for c in lst if c.get('pc') is not None]
        res[sig] = {
            'n_cursos': len(lst),
            'cpc_avg': round(sum(cpcs) / len(cpcs), 3) if cpcs else None,
            'cpc_max': max(cpcs) if cpcs else None,
            'ufs': sorted(set(c['u'] for c in lst if c.get('u'))),
        }
    return res

# ---------- main ----------

def main():
    print(f'Lendo de {IN_DIR}')

    cursos_17 = extract_cpc_2017(os.path.join(IN_DIR, 'CPC_2017.xlsx'))
    cursos_21 = extract_cpc_2021(os.path.join(IN_DIR, 'CPC_2021.xlsx'))
    print(f'CPC 2017 Física: {len(cursos_17)} cursos')
    print(f'CPC 2021 Física: {len(cursos_21)} cursos')

    igc17 = extract_igc(os.path.join(IN_DIR, 'IGC_2017.xlsx'), 2017, 'igc')
    igc21 = extract_igc(os.path.join(IN_DIR, 'IGC_2021.xlsx'), 2021, 'igc')
    print(f'IGC 2017: {len(igc17)} IES')
    print(f'IGC 2021: {len(igc21)} IES')

    # Anexar IGC ao curso
    for c in cursos_17:
        ig = igc17.get(c.get('ci'))
        if ig:
            c['igc_f'] = ig.get('igc_f')
            c['igc_c'] = ig.get('igc_c')
    for c in cursos_21:
        ig = igc21.get(c.get('ci'))
        if ig:
            c['igc_f'] = ig.get('igc_f')
            c['igc_c'] = ig.get('igc_c')

    cursos = cursos_17 + cursos_21

    # Limpar chaves com None para minificar
    cursos_min = []
    for c in cursos:
        cursos_min.append({k: v for k, v in c.items() if v not in (None, '')})

    ies_list = sorted(set(c['sg'] for c in cursos if c.get('sg')))

    dataset = {
        'metadata': {
            'fonte': 'INEP — Dados Abertos (Indicadores de Qualidade da Educação Superior)',
            'portal': 'https://www.gov.br/inep/pt-br/acesso-a-informacao/dados-abertos',
            'area': 'FÍSICA (LICENCIATURA + BACHARELADO)',
            'ciclos': [2017, 2021],
            'n_cursos': len(cursos_min),
            'n_ies': len(ies_list),
            'ufs': sorted(set(c['u'] for c in cursos if c.get('u'))),
            'gerado_em': __import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        },
        'legend': {
            'y': 'ano do ciclo ENADE',
            'ar': 'area de avaliacao',
            'g': 'grau academico',
            'ci': 'codigo IES',
            'ie': 'nome IES',
            'sg': 'sigla IES',
            'o': 'organizacao academica',
            'ct': 'categoria administrativa',
            'cc': 'codigo do curso (e-MEC)',
            'm': 'modalidade',
            'mu': 'municipio',
            'u': 'UF',
            'ni': 'n inscritos',
            'np': 'n participantes',
            'fg': 'nota bruta FG',
            'ce': 'nota bruta CE',
            'e':  'ENADE continuo',
            'ib': 'IDD bruta',
            'ip': 'IDD padronizada',
            'op': 'org didatico-pedagogica (padronizada)',
            'nf': 'infraestrutura (padronizada)',
            'of': 'oportunidade ampliacao formacao (padronizada)',
            'nd': 'n docentes',
            'ms': '% mestres (padronizada)',
            'dr': '% doutores (padronizada)',
            'rg': 'regime trabalho (padronizada)',
            'pc': 'CPC continuo',
            'pf': 'CPC faixa (1-5)',
            'igc_f': 'IGC da IES faixa (1-5)',
            'igc_c': 'IGC da IES continuo',
        },
        'ies_list': ies_list,
        'cursos': cursos_min,
    }

    # Gerar JSONs
    out_raiz = os.path.join(BASE, 'dados_inep.json')
    out_docs = os.path.join(BASE, 'docs', 'dados_inep.json')
    with open(out_raiz, 'w', encoding='utf-8') as f:
        json.dump(dataset, f, ensure_ascii=False, separators=(',', ':'))
    os.makedirs(os.path.dirname(out_docs), exist_ok=True)
    with open(out_docs, 'w', encoding='utf-8') as f:
        json.dump(dataset, f, ensure_ascii=False, separators=(',', ':'))

    raiz_size = os.path.getsize(out_raiz)
    print(f'\n== Dataset gerado ==')
    print(f'{out_raiz}: {raiz_size/1024:.1f} KB')
    print(f'{out_docs}: {raiz_size/1024:.1f} KB (copia)')
    print(f'\nResumo:')
    print(f'  Total cursos: {dataset["metadata"]["n_cursos"]}')
    print(f'  IES distintas: {dataset["metadata"]["n_ies"]}')
    print(f'  UFs cobertas: {len(dataset["metadata"]["ufs"])}')
    # Breakdown
    lic = [c for c in cursos_min if c.get('g') == 'Licenciatura']
    bac = [c for c in cursos_min if c.get('g') == 'Bacharelado']
    print(f'  Licenciatura: {len(lic)}')
    print(f'  Bacharelado: {len(bac)}')
    return dataset

if __name__ == '__main__':
    main()
